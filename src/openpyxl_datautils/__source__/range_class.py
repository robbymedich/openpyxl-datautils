from __future__ import annotations
from typing import Generator
from typing import Any
from contextlib import contextmanager
from itertools import islice
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import CellRange as FloatingCellRange
import openpyxl.styles as style
import openpyxl
import pandas as pd
import numpy as np
import re


@contextmanager
def load_workbook(file_path: str, read_only: bool = True, data_only: bool = True, **kwargs):
    """
    Wrapper function around openpyxl.load_workbook(). This function opens a workbook using a context manager, which
    helps to prevent locks on a file if an error is thrown in the program.

    Example usage:
    with load_workbook('my_workbook.xlsx') as workbook:
        df = CellRange.from_string(workbook, 'Sheet1!A1:A1').create_df(expand_range=True)

    Parameters:
         file_path: Location of the Excel workbook to load into memory
         read_only: Indicates if the file is opened in 'Read Only' mode. This greatly improves performance loading data.
         data_only: Controls whether cells with formulae have either the formula or the value stored the last time Excel
         read the worksheet (default).

    Returns:
        openpyxl Workbook object.
    """
    workbook = Workbook()
    try:
        workbook = openpyxl.load_workbook(filename=file_path, read_only=read_only, data_only=data_only, **kwargs)
        yield workbook
    finally:
        workbook.close()


class Immutable:
    """ Create a basic class to set immutable attributes on an object """
    _immutable_attr = tuple()

    def __setattr__(self, key, value):
        """ Prevent attributes in _immutable_attr from being set after __init__ """
        if key in self._immutable_attr and hasattr(self, key):
            raise AttributeError(f"{key} cannot be modified after the object's creation.")
        super().__setattr__(key, value)

    def __delattr__(self, item):
        """ Prevent attributes in _immutable_attr from being deleted """
        if item in self._immutable_attr:
            raise AttributeError(f"{item} cannot be deleted from the object.")
        super().__delattr__(item)


class WorksheetBoundaryError(Exception):
    """ Create a custom exception for any attempts to use cells outside the worksheet range """
    def __init__(self, worksheet: Worksheet):
        sheet_range = CellRange(
            worksheet=worksheet,
            start_row=worksheet.min_row,
            start_column=worksheet.min_column,
            end_row=worksheet.max_row,
            end_column=worksheet.max_column
        )
        self.message = f'Attempted to read values outside the used range of the worksheet.' \
                       f'Cells must be within {sheet_range}.'
        super().__init__(self.message)


class CellRange(Immutable):
    """
    Represents a range of cells within a Worksheet. Cells within the range can be read or written to as a group.

    Attributes
    ----------
    worksheet: openpyxl.Worksheet
        Excel worksheet containing the range of cells
    address: str
        Range address for the group of cells, ex. Sheet1!$A$1:$D$5
    shape: tuple
        Tuple with the number of rows and columns inside the range
    bounds: dict
        Dictionary with the starting and ending row and column index values
    cells: tuple[tuple[Cell]]
        Returns all cells in the range.
    values: Generator[tuple[Any]]
        Iterates over all rows in the range to return the values of each cell

    Methods
    -------
    from_string: CellRange
        Create a CellRange from a range address instead of indices for the range boundary
    from_worksheet: CellRange
        Create a CellRange object to represent ALL cells in the used range of a worksheet.
    current_region: CellRange
        Expand the range to the right, then down to include adjacent data (mimics Excel VBA's current_region method)
    create_df: pd.DataFrame
        Create a Pandas DataFrame with the data contained in a given CellRange
    write_df:
        Write data from a set of tuples to the values of a CellRange
    """
    __slots__ = ('worksheet', '_start_row', '_start_column', '_end_row', '_end_column', 'address',
                 '_end_unknown', '_FloatingCellRange')
    _immutable_attr = ('worksheet', '_start_row', '_start_column', '_end_row', '_end_column',
                       'address', '_FloatingCellRange')

    def __init__(self, worksheet: Worksheet, start_row: int, start_column: int,
                 end_row: int = None, end_column: int = None):
        """
        Create a CellRange object to represent a group of cells within a worksheet.

        Parameters:
             worksheet: openpyxl.Worksheet containing the range of cells
             start_row: starting row for the range
             start_column: starting column for the range
             end_row: ending row for the range
             end_column: ending column for the range
        """
        self._end_unknown = False
        self.worksheet = worksheet
        self._start_row = start_row
        self._start_column = start_column

        if end_row is None or end_column is None:
            # make sure both are None
            if end_row is not None or end_column is not None:
                raise ValueError(
                    'Invalid range, if either end_row or end_column is not passed, both must not be passed.')

            end_row, end_column = start_row, start_column

        self._end_row = end_row
        self._end_column = end_column

        # coordinate_to_tuple and openpyxl.CellRange validate range is valid
        # get string representation of the range address
        start_cell_str = f'${openpyxl.utils.get_column_letter(self._start_column)}${self._start_row}'
        end_cell_str = f'${openpyxl.utils.get_column_letter(self._end_column)}${self._end_row}'
        sheet_title = self.worksheet.title
        sheet_title = f"'{sheet_title}'" if re.search(r'\s', sheet_title) else sheet_title

        self.address = f'{sheet_title}!{start_cell_str}:{end_cell_str}'
        self._FloatingCellRange = FloatingCellRange(self.address)

    def __setattr__(self, key, value):
        """ Allow cell values to be set within the range """
        if key == 'values':
            self._set_cell_values(cell_values=value)
        else:
            super().__setattr__(key, value)

    def __repr__(self):
        """ String representation should be the range address """
        return self.address

    def __eq__(self, other):
        """ Allow '==' comparison checks """
        if isinstance(other, CellRange):
            return self.address == other.address
        return False

    @classmethod
    def from_string(cls, parent: Worksheet | Workbook, range_address: str) -> CellRange:
        """
        Create a CellRange object to represent a group of cells within a worksheet.

        Parameters:
             parent: either a openpyxl.Worksheet or openpyxl.Workbook containing the range of cells
             range_address: address for the group of cells to use when creating the CellRange
        """
        if isinstance(parent, str):
            raise TypeError("Expected a 'Worksheet' or 'Workbook' object. Received 'str'")

        if isinstance(parent, Workbook):
            try:
                sheet_name, range_address = range_address.split('!')
            except ValueError:
                if range_address.count('!') == 0:
                    raise ValueError(f'{range_address} must contain the sheetname when the parent is a Workbook')
                raise ValueError(f'{range_address} is not a valid Excel Range.')

            worksheet = parent[sheet_name]

        else:
            if '!' in range_address:
                try:
                    sheet_name, range_address = range_address.split('!')
                except ValueError:
                    raise ValueError(f'{range_address} is not a valid Excel Range.')

                if sheet_name != parent.title:
                    raise ValueError(f'worksheet in {range_address} does not refer to parent Worksheet supplied.')
            worksheet = parent

        start_column, start_row, end_column, end_row = openpyxl.utils.range_boundaries(range_address)
        return cls(worksheet, start_row, start_column, end_row, end_column)

    @classmethod
    def from_worksheet(cls, worksheet: Worksheet) -> CellRange:
        """
        Create a CellRange object to represent ALL cells in the used range of a worksheet.

        Parameters:
             worksheet: openpyxl.Worksheet containing the range of cells
        """
        return cls(
            worksheet=worksheet,
            start_row=worksheet.min_row,
            start_column=worksheet.min_column,
            end_row=worksheet.max_row,
            end_column=worksheet.max_column,
        )

    @property
    def bounds(self) -> dict:
        """ Dictionary with the starting and ending row and column index values """
        return {
            'start_row': self._start_row,
            'start_column': self._start_column,
            'end_row': self._end_row,
            'end_column': self._end_column,
        }

    @property
    def shape(self) -> tuple:
        """ Tuple with the number of rows and columns inside the range """
        return (
            self._end_row - self._start_row + 1,
            self._end_column - self._start_column + 1,
        )

    @property
    def cells(self) -> tuple[tuple[Cell]]:
        """ Returns all rows and cells in the CellRange object -> tuple[tuple[Cell]] """
        return self.worksheet[self.address.split('!')[1]]

    @property
    def values(self) -> Generator[tuple[Any]]:
        """ Iterates over all rows in the range to return the values of each cell -> Generator[tuple[Any]] """
        # Make sure range is within the used area on the worksheet
        if self._start_row < self.worksheet.min_row or self._start_column < self.worksheet.min_column or \
                self._end_row > self.worksheet.max_row or self._end_column > self.worksheet.max_column:
            raise WorksheetBoundaryError(self.worksheet)

        for i, row_values in enumerate(self.worksheet.values):
            if i < self._start_row - 1:
                continue
            if self._end_unknown is False and i > self._end_row - 1:
                break

            row_values = tuple(islice(row_values, self._start_column - 1, self._end_column))
            if self._end_unknown is True and any(row_values) is False:
                break
            yield row_values

    def _set_cell_values(self, cell_values: tuple[tuple[Any]] | np.ndarray):
        """ Write data from a set of tuples to the values of a CellRange """
        # Create array with new cell values
        updated_values = np.array(cell_values)
        if len(updated_values.shape) != 2:
            raise ValueError(f'Values provided created a {len(updated_values.shape)}-D Array instead of a 2-D Array.\n'
                             f'To create a proper array, cell_values should be a form similar to tuple[tuple[Any]].')

        # Make sure the shape of the array matches the range
        if self.shape != updated_values.shape:
            raise ValueError(f'Cannot set {updated_values.shape} values in a range which is {self.shape}.')

        # Set all values to the range
        for i, row in enumerate(self.cells):
            for j, cell in enumerate(row):
                cell.value = updated_values[i, j]

    def _get_last_column(self) -> int:
        """ Get the last column with data from a given starting point """
        start_cell = self.worksheet.cell(self._start_row, self._start_column)
        current_cell = start_cell
        col_index = start_cell.column

        while current_cell.value is not None:
            col_index += 1
            current_cell = self.worksheet.cell(self._start_row, col_index)

        return max(col_index - 1, self._end_column)

    def _get_last_row(self) -> int:
        """ Get the last row with data from a given starting point """
        end_unknown = self._end_unknown
        self._end_unknown = True

        i = 0
        for i, _ in enumerate(self.values):
            pass

        self._end_unknown = end_unknown
        return max(i + self._start_row, self._end_row)

    def current_region(self) -> CellRange:
        """
        Expand the range to the right, then down to include adjacent data ("mimics" Excel VBA's current_region method).

        Returns:
            CellRange object containing adjacent data from the source CellRange.
        """
        cell_range = CellRange(
            worksheet=self.worksheet,
            start_row=self._start_row,
            start_column=self._start_column,
            end_row=self._start_row,
            end_column=self._get_last_column(),
        )
        return CellRange(
            worksheet=cell_range.worksheet,
            start_row=cell_range._start_row,
            start_column=cell_range._start_column,
            end_row=cell_range._get_last_row(),
            end_column=cell_range._end_column,
        )

    def create_df(self, has_headers: bool = True, expand_range: bool = False) -> pd.DataFrame:
        """
        Create a Pandas DataFrame with the data contained in a given CellRange.

        Parameters:
            has_headers: Indicates if the first row of data represents column names for the dataframe
            expand_range: Indicates if the range should automatically be expanded to include adjacent data.
                - If the range is one cell, the range will be expanded to the right, then down
                - If the range is one row, the range will be expanded down

        Returns:
            DataFrame with the data in the CellRange
        """
        end_unknown = self._end_unknown
        cell_range = self
        if expand_range:
            # make sure range is a single cell or single row
            if self._start_row != self._end_row:
                raise ValueError('Range can only be expanded if the original range refers to one row or one cell.')

            if self._start_column == self._end_column:
                cell_range = CellRange(
                    worksheet=self.worksheet,
                    start_row=self._start_row,
                    start_column=self._start_column,
                    end_row=self._end_row,
                    end_column=self._get_last_column(),
                )
                cell_range._end_unknown = True
                row_values = cell_range.values
            else:
                self._end_unknown = True
                row_values = self.values
        else:
            row_values = self.values

        if has_headers is False:
            header_range = CellRange(
                worksheet=cell_range.worksheet,
                start_row=cell_range._start_row,
                start_column=cell_range._start_column,
                end_row=cell_range._start_row,
                end_column=cell_range._end_column,
            )
            columns = [cell.column_letter for cell in header_range.cells[0]]
        else:
            columns = next(row_values)  # first row is column names

        df = pd.DataFrame(data=row_values, columns=columns)
        start_row = cell_range._start_row + 1 if has_headers else cell_range._start_row
        df.index = np.arange(start_row, start_row + df.shape[0])

        self._end_unknown = end_unknown
        return df

    def write_df(self, df: pd.DataFrame, expand_range: bool = False, apply_style: bool = True,
                 adjust_widths: bool = True):
        """
        Write a dataframe to the CellRange and optionally apply formatting to mimic df.to_excel()

        Parameters:
            df: Pandas DataFrame to write to the CellRange
            expand_range: Indicates if the range should automatically be expanded to include adjacent data.
                - If the range is one cell, the range will be expanded to the right, then down
                - If the range is one row, the range will be expanded down
            apply_style: Indicates if style from df.to_excel() should be matched when writing data to the sheet.
            adjust_widths: Adjust column widths based on the number of characters in the source column
        """
        # Get data
        updated_values = np.vstack((df.columns.values, df.values))

        # Get write range
        if expand_range is True:
            # make sure range is a single cell or single row
            if self._start_row != self._end_row:
                raise ValueError('Range can only be expanded if the original range refers to one row or one cell.')

            if self._start_column == self._end_column:
                cell_range = CellRange(
                    worksheet=self.worksheet,
                    start_row=self._start_row,
                    start_column=self._start_column,
                    end_row=self._start_row + updated_values.shape[0] - 1,
                    end_column=self._start_column + updated_values.shape[1] - 1,
                )
            else:
                cell_range = CellRange(
                    worksheet=self.worksheet,
                    start_row=self._start_row,
                    start_column=self._start_column,
                    end_row=self._start_row + updated_values.shape[0] - 1,
                    end_column=self._end_column,
                )
        else:
            cell_range = self

        # Write values to the worksheet
        cell_range._set_cell_values(updated_values)

        # Apply formatting
        if not apply_style:
            return
        _style_headers(CellRange(
            worksheet=cell_range.worksheet,
            start_row=cell_range._start_row,
            start_column=cell_range._start_column,
            end_row=cell_range._start_row,
            end_column=cell_range._end_column,
        ))

        # Estimate column widths by the number of characters in the df
        if not adjust_widths:
            return
        for j, col in enumerate(df):
            max_len = max(df[col].astype(str).apply(len).max(), len(col))

            col_letter = openpyxl.utils.get_column_letter(cell_range._start_column + j)
            cell_range.worksheet.column_dimensions[col_letter].width = max_len


def _style_headers(cell_range: CellRange):
    """
    Style the header row of a data set in Excel to have bold centered text with a border around each cell.
    """
    thin_border = style.Side(style='thin')

    for cell in cell_range.cells[0]:
        cell.font = style.Font(bold=True)
        cell.alignment = style.Alignment(horizontal='center')
        cell.border = style.Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)
