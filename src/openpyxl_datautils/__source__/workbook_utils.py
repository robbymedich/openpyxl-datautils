from __future__ import annotations
from contextlib import contextmanager
from openpyxl.workbook.workbook import Workbook
import openpyxl


@contextmanager
def load_workbook(file_path: str, read_only: bool = True, data_only: bool = True, **kwargs):
    """
    Wrapper function around openpyxl.load_workbook(). This function opens a workbook using a context manager, which
    helps to prevent locks on a file if an error is thrown in the program.

    Example usage:
    with load_workbook('my_workbook.xlsx') as workbook:
        df = CellRange.from_string(workbook, 'Sheet1!A1:A1').create_df(expand_range=True)

    Parameters:
         file_path: Location of the Excel workbook to load into memory
         read_only: Indicates if the file is opened in 'Read Only' mode. This greatly improves performance loading data.
         data_only: Controls whether cells with formulae have either the formula or the value stored the last time Excel
         read the worksheet (default).

    Returns:
        openpyxl Workbook object.
    """
    workbook = Workbook()
    try:
        workbook = openpyxl.load_workbook(filename=file_path, read_only=read_only, data_only=data_only, **kwargs)
        yield workbook
    finally:
        workbook.close()
