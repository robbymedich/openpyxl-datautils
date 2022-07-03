from __future__ import annotations
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from itertools import islice
import pandas as pd


class Immutable:
    """ Create a basic class to set immutable attributes on an object """
    _immutable_attr = tuple()

    def __setattr__(self, key, value):
        """ Prevent attributes in _immutable_attr from being set after __init__ """
        if key in self._immutable_attr and hasattr(self, key):
            raise AttributeError(f"{key} cannot be modified after the object's creation.")
        super().__setattr__(key, value)

    def __delattr__(self, item):
        """ Prevent attributes in _immutable_attr from being deleted """
        if item in self._immutable_attr:
            raise AttributeError(f"{item} cannot be deleted from the object.")
        super().__delattr__(item)


class SheetBoundaryError(Exception):
    """ Create a custom exception for any attempts to use cells outside the sheet range """
    def __init__(self, sheet: Worksheet):
        self.sheet = sheet

    def __str__(self):
        sheet_range = ExcelRange(
            sheet=self.sheet,
            start_row=self.sheet.min_row,
            start_column=self.sheet.min_column,
            end_row=self.sheet.max_row,
            end_column=self.sheet.max_column
        )
        return f'Attempted to read values outside the used range of the sheet. Cells must be within {sheet_range}.'


class ExcelRange(Immutable):
    """
    Must add checks to ensure end cell is greater then start cell
    Add functionality to set values in the range
    """
    __slots__ = ('sheet', '_start_row', '_start_column', '_end_row', '_end_column', 'address', '_end_unknown')
    _immutable_attr = ('sheet', '_start_row', '_start_column', '_end_row', '_end_column', 'address')

    def __init__(self, sheet: Worksheet, start_row: int, start_column: int,
                 end_row: int = None, end_column: int = None):
        self._end_unknown = False
        self.sheet = sheet
        self._start_row = start_row
        self._start_column = start_column

        if end_row is None or end_column is None:
            # make sure both are None
            if end_row is not None or end_column is not None:
                raise ValueError(
                    'Invalid range, if either end_row or end_column is not passed, both must not be passed.')

            end_row, end_column = start_row, start_column

        self._end_row = end_row
        self._end_column = end_column

        # make sure range is valid by calling coordinate_to_tuple on both cells
        try:
            self.address = self.__repr__()
        except Exception as exception:
            raise exception

        # do not allow zeros - this is not caught by openpyxl
        if any(boundary == 0 for boundary in (self._start_row, self._start_column, self._end_row, self._end_column)):
            raise ValueError(f'{self.address} is not a valid Excel Range.')

    def __setattr__(self, key, value):
        if key == 'values':
            raise NotImplementedError('Setting values on a range has not yet been implemented')
        super().__setattr__(key, value)

    def __repr__(self):
        start_cell_str = f'${openpyxl.utils.get_column_letter(self._start_column)}${self._start_row}'
        end_cell_str = f'${openpyxl.utils.get_column_letter(self._end_column)}${self._end_row}'
        return f'{self.sheet.title}!{start_cell_str}:{end_cell_str}'

    def __str__(self):
        return self.__repr__()

    def __eq__(self, other):
        if isinstance(other, ExcelRange):
            return self.address == other.address
        return False

    @classmethod
    def from_string(cls, parent: Worksheet | Workbook, range_address: str):
        if isinstance(parent, Workbook):
            try:
                sheet_name, range_address = range_address.split('!')
            except ValueError:
                if range_address.count('!') == 0:
                    raise ValueError(f'{range_address} must contain the sheetname when the parent is a Workbook')
                raise ValueError(f'{range_address} is not a valid Excel Range.')

            sheet = parent[sheet_name]

        else:
            if '!' in range_address:
                try:
                    sheet_name, range_address = range_address.split('!')
                except ValueError:
                    raise ValueError(f'{range_address} is not a valid Excel Range.')

                if sheet_name != parent.title:
                    raise ValueError(f'Sheet in {range_address} does not refer to parent Worksheet supplied.')
            sheet = parent

        start_column, start_row, end_column, end_row = openpyxl.utils.range_boundaries(range_address)
        return cls(sheet, start_row, start_column, end_row, end_column)

    @property
    def cells(self):
        return self.sheet[self.address.split('!')[1]]

    @property
    def values(self):
        # Make sure range is within the used area on the sheet
        if self._start_row < self.sheet.min_row or self._start_column < self.sheet.min_column or \
                self._end_row > self.sheet.max_row or self._end_column > self.sheet.max_column:
            raise SheetBoundaryError(self.sheet)

        row_start = self._start_row - 1
        row_end = self._end_row - 1
        col_start = self._start_column + (self._start_column - self.sheet.min_column) - 1
        col_end = col_start + (self._end_column - self._start_column) + 1

        for i, row_values in enumerate(self.sheet.values):
            if i < row_start:
                continue
            if self._end_unknown is False and i > row_end:
                break

            row_values = tuple(islice(row_values, col_start, col_end))
            if self._end_unknown is True and any(row_values) is False:
                break
            yield row_values

    def _get_last_column(self):
        start_cell = self.sheet.cell(self._start_row, self._start_column)
        current_cell = start_cell
        col_index = start_cell.column

        while current_cell.value is not None:
            col_index += 1
            current_cell = self.sheet.cell(self._start_row, col_index)

        return max(col_index - 1, self._end_column)

    def _get_last_row(self):
        end_unknown = self._end_unknown
        self._end_unknown = True

        i = 0
        for i, _ in enumerate(self.values):
            pass

        self._end_unknown = end_unknown
        return max(i + self._start_row, self._end_row)

    def create_df(self, expand_range: bool = False):
        if expand_range:
            # make sure range is a single cell
            if (self._start_row, self._start_column) != (self._end_row, self._end_column):
                raise ValueError('Range can only be expanded if the original range refers to one cell.')

            excel_range = ExcelRange(
                sheet=self.sheet,
                start_row=self._start_row,
                start_column=self._start_column,
                end_row=self._end_row,
                end_column=self._get_last_column()
            )
            excel_range._end_unknown = True
            row_values = excel_range.values
        else:
            row_values = self.values

        columns = next(row_values)  # first row should be column names
        return pd.DataFrame(data=row_values, columns=columns)

    def current_region(self):
        excel_range = ExcelRange(
            sheet=self.sheet,
            start_row=self._start_row,
            start_column=self._start_column,
            end_row=self._start_row,
            end_column=self._get_last_column(),
        )
        return ExcelRange(
            sheet=excel_range.sheet,
            start_row=excel_range._start_row,
            start_column=excel_range._start_column,
            end_row=excel_range._get_last_row(),
            end_column=excel_range._end_column,
        )
